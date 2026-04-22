#!/usr/bin/env python3

import argparse
import json
import os
import subprocess
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class OpenShiftHealthCheckApp:
    def __init__(self, namespace=None, label_selector=None, output_dir="reports", report_name=None):
        self.namespace = namespace
        self.label_selector = label_selector
        self.output_dir = output_dir
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.report_name = report_name or f"openshift-health-report-{self.timestamp}.xlsx"
        self.report_path = os.path.join(self.output_dir, self.report_name)
        self.results = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "namespace": namespace or "all-namespaces",
                "label_selector": label_selector or "none"
            },
            "summary": {},
            "cluster": {},
            "nodes": [],
            "pods": [],
            "deployments": [],
            "services": [],
            "routes": [],
            "events": [],
            "probes": []
        }

    def run_oc(self, args):
        command = ["oc"] + args
        completed = subprocess.run(command, capture_output=True, text=True)
        return completed.returncode, completed.stdout.strip(), completed.stderr.strip()

    def ensure_requirements(self):
        if not shutil.which("oc"):
            raise RuntimeError("OpenShift CLI 'oc' is not installed or not available in PATH.")

        code, output, _ = self.run_oc(["whoami"])
        if code != 0:
            raise RuntimeError("Not logged in to OpenShift. Run: oc login <cluster-url>")
        self.results["cluster"]["current_user"] = output

    def collect_cluster_info(self):
        code, output, _ = self.run_oc(["version", "-o", "json"])
        if code == 0 and output:
            try:
                payload = json.loads(output)
                self.results["cluster"]["client_version"] = payload.get("clientVersion", {}).get("gitVersion", "unknown")
                self.results["cluster"]["server_version"] = payload.get("openshiftVersion", "unknown")
            except json.JSONDecodeError:
                self.results["cluster"]["server_version"] = "unknown"
        else:
            self.results["cluster"]["server_version"] = "unknown"

    def collect_nodes(self):
        code, output, _ = self.run_oc(["get", "nodes", "-o", "json"])
        if code != 0 or not output:
            return

        payload = json.loads(output)
        for item in payload.get("items", []):
            conditions = item.get("status", {}).get("conditions", [])
            ready_status = "Unknown"
            for condition in conditions:
                if condition.get("type") == "Ready":
                    ready_status = "Ready" if condition.get("status") == "True" else "NotReady"
                    break

            capacity = item.get("status", {}).get("capacity", {})
            labels = item.get("metadata", {}).get("labels", {})
            roles = sorted({
                key.split("/")[-1]
                for key in labels.keys()
                if key.startswith("node-role.kubernetes.io/")
            })

            self.results["nodes"].append({
                "name": item.get("metadata", {}).get("name", ""),
                "status": ready_status,
                "roles": ", ".join(roles) if roles else "worker",
                "cpu": capacity.get("cpu", ""),
                "memory": capacity.get("memory", ""),
                "os": item.get("status", {}).get("nodeInfo", {}).get("operatingSystem", ""),
                "kernel": item.get("status", {}).get("nodeInfo", {}).get("kernelVersion", "")
            })

    def collect_pods(self):
        args = ["get", "pods", "-o", "json"]
        if self.namespace:
            args.extend(["-n", self.namespace])
        else:
            args.append("-A")
        if self.label_selector:
            args.extend(["-l", self.label_selector])

        code, output, _ = self.run_oc(args)
        if code != 0 or not output:
            return

        payload = json.loads(output)
        for item in payload.get("items", []):
            metadata = item.get("metadata", {})
            status = item.get("status", {})
            spec = item.get("spec", {})
            namespace = metadata.get("namespace", "")
            pod_name = metadata.get("name", "")
            phase = status.get("phase", "Unknown")
            node_name = spec.get("nodeName", "")
            container_statuses = status.get("containerStatuses", []) or []

            ready_count = 0
            restart_count = 0
            waiting_reasons = []
            terminated_reasons = []

            for container in container_statuses:
                if container.get("ready"):
                    ready_count += 1
                restart_count += container.get("restartCount", 0)

                waiting_state = container.get("state", {}).get("waiting")
                terminated_state = container.get("state", {}).get("terminated")

                if waiting_state and waiting_state.get("reason"):
                    waiting_reasons.append(waiting_state.get("reason"))
                if terminated_state and terminated_state.get("reason"):
                    terminated_reasons.append(terminated_state.get("reason"))

            reason = ", ".join(waiting_reasons or terminated_reasons) or status.get("reason", "")
            self.results["pods"].append({
                "namespace": namespace,
                "name": pod_name,
                "phase": phase,
                "ready": f"{ready_count}/{len(container_statuses)}",
                "restarts": restart_count,
                "node": node_name,
                "reason": reason or "OK"
            })

    def collect_deployments(self):
        args = ["get", "deployments", "-o", "json"]
        if self.namespace:
            args.extend(["-n", self.namespace])
        else:
            args.append("-A")
        if self.label_selector:
            args.extend(["-l", self.label_selector])

        code, output, _ = self.run_oc(args)
        if code != 0 or not output:
            return

        payload = json.loads(output)
        for item in payload.get("items", []):
            metadata = item.get("metadata", {})
            spec = item.get("spec", {})
            status = item.get("status", {})
            desired = spec.get("replicas", 0)
            available = status.get("availableReplicas", 0)
            ready = status.get("readyReplicas", 0)

            self.results["deployments"].append({
                "namespace": metadata.get("namespace", ""),
                "name": metadata.get("name", ""),
                "desired": desired,
                "ready": ready,
                "available": available,
                "health": "Healthy" if desired == available else "Review"
            })

    def collect_services(self):
        args = ["get", "svc", "-o", "json"]
        if self.namespace:
            args.extend(["-n", self.namespace])
        else:
            args.append("-A")
        if self.label_selector:
            args.extend(["-l", self.label_selector])

        code, output, _ = self.run_oc(args)
        if code != 0 or not output:
            return

        payload = json.loads(output)
        for item in payload.get("items", []):
            metadata = item.get("metadata", {})
            spec = item.get("spec", {})
            ports = spec.get("ports", [])
            port_text = ", ".join(str(port.get("port", "")) for port in ports)

            self.results["services"].append({
                "namespace": metadata.get("namespace", ""),
                "name": metadata.get("name", ""),
                "type": spec.get("type", ""),
                "cluster_ip": spec.get("clusterIP", ""),
                "ports": port_text
            })

    def collect_routes(self):
        args = ["get", "routes.route.openshift.io", "-o", "json"]
        if self.namespace:
            args.extend(["-n", self.namespace])
        else:
            args.append("-A")
        if self.label_selector:
            args.extend(["-l", self.label_selector])

        code, output, _ = self.run_oc(args)
        if code != 0 or not output:
            return

        payload = json.loads(output)
        for item in payload.get("items", []):
            metadata = item.get("metadata", {})
            spec = item.get("spec", {})
            self.results["routes"].append({
                "namespace": metadata.get("namespace", ""),
                "name": metadata.get("name", ""),
                "host": spec.get("host", ""),
                "path": spec.get("path", ""),
                "service": spec.get("to", {}).get("name", "")
            })

    def collect_events(self):
        args = ["get", "events", "-o", "json"]
        if self.namespace:
            args.extend(["-n", self.namespace])
        else:
            args.append("-A")

        code, output, _ = self.run_oc(args)
        if code != 0 or not output:
            return

        payload = json.loads(output)
        items = payload.get("items", [])[-50:]
        for item in items:
            involved = item.get("involvedObject", {})
            self.results["events"].append({
                "namespace": item.get("metadata", {}).get("namespace", ""),
                "type": item.get("type", ""),
                "reason": item.get("reason", ""),
                "object": f"{involved.get('kind', '')}/{involved.get('name', '')}",
                "message": item.get("message", "")
            })

    def collect_probes(self):
        args = ["get", "deployments", "-o", "json"]
        if self.namespace:
            args.extend(["-n", self.namespace])
        else:
            args.append("-A")
        if self.label_selector:
            args.extend(["-l", self.label_selector])

        code, output, _ = self.run_oc(args)
        if code != 0 or not output:
            return

        payload = json.loads(output)
        for item in payload.get("items", []):
            namespace = item.get("metadata", {}).get("namespace", "")
            deployment_name = item.get("metadata", {}).get("name", "")
            containers = item.get("spec", {}).get("template", {}).get("spec", {}).get("containers", [])

            for container in containers:
                for probe_type in ["livenessProbe", "readinessProbe", "startupProbe"]:
                    probe = container.get(probe_type)
                    if not probe:
                        continue

                    http_get = probe.get("httpGet", {})
                    self.results["probes"].append({
                        "namespace": namespace,
                        "deployment": deployment_name,
                        "container": container.get("name", ""),
                        "probe_type": probe_type,
                        "path": http_get.get("path", ""),
                        "port": http_get.get("port", ""),
                        "initial_delay": probe.get("initialDelaySeconds", "")
                    })

    def build_summary(self):
        total_nodes = len(self.results["nodes"])
        ready_nodes = len([n for n in self.results["nodes"] if n["status"] == "Ready"])
        total_pods = len(self.results["pods"])
        running_pods = len([p for p in self.results["pods"] if p["phase"] == "Running"])
        unhealthy_pods = len([p for p in self.results["pods"] if p["phase"] != "Running"])
        review_deployments = len([d for d in self.results["deployments"] if d["health"] != "Healthy"])

        self.results["summary"] = {
            "nodes_ready": f"{ready_nodes}/{total_nodes}",
            "pods_running": f"{running_pods}/{total_pods}",
            "unhealthy_pods": unhealthy_pods,
            "deployments_needing_review": review_deployments,
            "services": len(self.results["services"]),
            "routes": len(self.results["routes"]),
            "events": len(self.results["events"])
        }

    def style_sheet(self, sheet, title):
        title_fill = PatternFill("solid", fgColor="0F243E")
        header_fill = PatternFill("solid", fgColor="1F4E78")
        alt_fill = PatternFill("solid", fgColor="F7FBFF")
        healthy_fill = PatternFill("solid", fgColor="E2F0D9")
        warning_fill = PatternFill("solid", fgColor="FFF2CC")
        critical_fill = PatternFill("solid", fgColor="FCE4D6")
        neutral_fill = PatternFill("solid", fgColor="F3F6F9")
        thin_border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )

        max_column = sheet.max_column
        if max_column > 0:
            sheet.insert_rows(1)
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
            title_cell = sheet.cell(row=1, column=1, value=f"OpenShift Health Check - {title}")
            title_cell.font = Font(bold=True, color="FFFFFF", size=14)
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.border = thin_border
            sheet.row_dimensions[1].height = 24

        header_row = 2
        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row_index, row in enumerate(sheet.iter_rows(min_row=3), start=3):
            row_fill = alt_fill if row_index % 2 == 1 else neutral_fill
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
                if cell.value is None or cell.value == "":
                    cell.fill = row_fill
                    continue

                value = str(cell.value)
                if value in ["Healthy", "Ready", "Running", "YES", "OK"]:
                    cell.fill = healthy_fill
                elif value in ["Review", "Pending", "Warning", "NotReady", "Unknown"]:
                    cell.fill = warning_fill
                elif value in ["Failed", "Error", "Critical"]:
                    cell.fill = critical_fill
                else:
                    cell.fill = row_fill

        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = sheet.dimensions

        for column_index in range(1, sheet.max_column + 1):
            visible_cells = []
            for row_index in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                if cell.value is not None:
                    visible_cells.append(cell)

            if not visible_cells:
                continue

            length = max(len(str(cell.value or "")) for cell in visible_cells)
            column_letter = get_column_letter(column_index)
            sheet.column_dimensions[column_letter].width = min(max(length + 3, 14), 50)

    def add_sheet(self, workbook, title, headers, rows):
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        if rows:
            for row in rows:
                sheet.append(row)
        else:
            sheet.append(["No data available"] + [""] * (len(headers) - 1))
        self.style_sheet(sheet, title)

    def generate_excel_report(self):
        os.makedirs(self.output_dir, exist_ok=True)
        workbook = Workbook()
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)

        workbook.properties.creator = "OpenShift Health Check Tool"
        workbook.properties.title = "OpenShift Health Check Report"
        workbook.properties.subject = "Cluster health workbook"
        workbook.properties.description = "Formatted OpenShift health report with multiple worksheets"

        self.add_sheet(
            workbook,
            "Summary",
            ["Metric", "Value"],
            [[key, value] for key, value in self.results["summary"].items()]
        )

        self.add_sheet(
            workbook,
            "Cluster",
            ["Field", "Value"],
            [[key, value] for key, value in self.results["cluster"].items()]
        )

        self.add_sheet(
            workbook,
            "Nodes",
            ["Name", "Status", "Roles", "CPU", "Memory", "OS", "Kernel"],
            [[n["name"], n["status"], n["roles"], n["cpu"], n["memory"], n["os"], n["kernel"]] for n in self.results["nodes"]]
        )

        self.add_sheet(
            workbook,
            "Pods",
            ["Namespace", "Pod", "Phase", "Ready", "Restarts", "Node", "Reason"],
            [[p["namespace"], p["name"], p["phase"], p["ready"], p["restarts"], p["node"], p["reason"]] for p in self.results["pods"]]
        )

        self.add_sheet(
            workbook,
            "Deployments",
            ["Namespace", "Deployment", "Desired", "Ready", "Available", "Health"],
            [[d["namespace"], d["name"], d["desired"], d["ready"], d["available"], d["health"]] for d in self.results["deployments"]]
        )

        self.add_sheet(
            workbook,
            "Services",
            ["Namespace", "Service", "Type", "ClusterIP", "Ports"],
            [[s["namespace"], s["name"], s["type"], s["cluster_ip"], s["ports"]] for s in self.results["services"]]
        )

        self.add_sheet(
            workbook,
            "Routes",
            ["Namespace", "Route", "Host", "Path", "Service"],
            [[r["namespace"], r["name"], r["host"], r["path"], r["service"]] for r in self.results["routes"]]
        )

        self.add_sheet(
            workbook,
            "Events",
            ["Namespace", "Type", "Reason", "Object", "Message"],
            [[e["namespace"], e["type"], e["reason"], e["object"], e["message"]] for e in self.results["events"]]
        )

        self.add_sheet(
            workbook,
            "Probes",
            ["Namespace", "Deployment", "Container", "Probe Type", "Path", "Port", "Initial Delay"],
            [[p["namespace"], p["deployment"], p["container"], p["probe_type"], p["path"], p["port"], p["initial_delay"]] for p in self.results["probes"]]
        )

        workbook.save(self.report_path)

    def run(self):
        print("Starting OpenShift health check...")
        self.ensure_requirements()
        self.collect_cluster_info()
        self.collect_nodes()
        self.collect_pods()
        self.collect_deployments()
        self.collect_services()
        self.collect_routes()
        self.collect_events()
        self.collect_probes()
        self.build_summary()
        self.generate_excel_report()
        print(f"Report generated: {self.report_path}")


def parse_args():
    parser = argparse.ArgumentParser(description="Generate an OpenShift health check Excel report.")
    parser.add_argument("--namespace", help="Target namespace. Default is all namespaces.")
    parser.add_argument("--label-selector", help="Optional label selector, for example app=myapp")
    parser.add_argument("--output-dir", default="reports", help="Directory for generated reports")
    parser.add_argument("--report-name", help="Custom report filename")
    return parser.parse_args()


def main():
    args = parse_args()
    app = OpenShiftHealthCheckApp(
        namespace=args.namespace,
        label_selector=args.label_selector,
        output_dir=args.output_dir,
        report_name=args.report_name
    )
    app.run()


if __name__ == "__main__":
    import shutil
    main()
